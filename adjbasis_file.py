import json
import csv
import openpyxl
import math
from openpyxl import Workbook

# диапазоны высот опор
RANGES = {
    "BS1" : (32, 52),
    "BS2" : (52, 69),
    "BS3" : (69, 103),
    "BS4" : (103, 134),
    "BS5" : (134, 217),
    "BS6" : (217, 331)
}

# чтение координат точек и треугольников из файла Excel
# Х и Y в метрах, Z в миллиметрах
POINTS = {}
wb = openpyxl.load_workbook("indata.xlsx", data_only=True)  
sheet = wb["POINTS"]
print(sheet.max_row)
for r in range(2, sheet.max_row+1):
    POINTS[r-1] = {"X" : sheet.cell(r, 2).value, "Y" : sheet.cell(r, 3).value, "Z" : sheet.cell(r, 4).value}
print(POINTS)  

TRIANGLES = {}
sheet = wb["TRIANGLES"]
print(sheet.max_row)
for r in range(2, sheet.max_row+1):
    TRIANGLES[r-1] = {"A" : POINTS[sheet.cell(r, 2).value], "B" : POINTS[sheet.cell(r, 3).value], "C" : POINTS[sheet.cell(r, 4).value]}
print(TRIANGLES)          

STEPX = 0.6 # шаг по оси Х в метрах
STEPY = 0.4 # шаг по оси Y в метрах

MAXX = -1000
MAXY = -1000

for k, v in POINTS.items():
    if v["X"] > MAXX:
        MAXX = v["X"]
    if v["Y"] > MAXY:
        MAXY = v["Y"]

print(MAXX)
print(MAXY)        
#MAXX = 19.2
#MAXY = 7.6

# определяет принадлежит ли точка (x, y) треугольнику TR
def inTriangle(x, y, TR):
    a = (TR["A"]["X"] - x) * (TR["B"]["Y"] - TR["A"]["Y"]) - (TR["B"]["X"] - TR["A"]["X"]) * (TR["A"]["Y"] - y)
    b = (TR["B"]["X"] - x) * (TR["C"]["Y"] - TR["B"]["Y"]) - (TR["C"]["X"] - TR["B"]["X"]) * (TR["B"]["Y"] - y)
    c = (TR["C"]["X"] - x) * (TR["A"]["Y"] - TR["C"]["Y"]) - (TR["A"]["X"] - TR["C"]["X"]) * (TR["C"]["Y"] - y)
 
    if (a >= 0 and b >= 0 and c >= 0) or (a <= 0 and b <= 0 and c <= 0):
        return True

# вычисляет высоту опоры в точке (x, y), предварительно определяя какому треугольнику она принадлежит
def Z(x, y):
    for key, T in TRIANGLES.items():
        if inTriangle(x, y, T):
            detA = (T["B"]["Y"] - T["A"]["Y"]) * (T["C"]["Z"] - T["A"]["Z"]) - (T["C"]["Y"] - T["A"]["Y"]) * (T["B"]["Z"] - T["A"]["Z"])
            detB = (T["B"]["X"] - T["A"]["X"]) * (T["C"]["Z"] - T["A"]["Z"]) - (T["C"]["X"] - T["A"]["X"]) * (T["B"]["Z"] - T["A"]["Z"])
            detC = (T["B"]["X"] - T["A"]["X"]) * (T["C"]["Y"] - T["A"]["Y"]) - (T["C"]["X"] - T["A"]["X"]) * (T["B"]["Y"] - T["A"]["Y"])
            return ((y - T["A"]["Y"]) * detB - (x - T["A"]["X"]) * detA) / detC + T["A"]["Z"]                      
    return 0

SUPPS = [] # список координат и высот всех опор на участке
SUPTYPES = {"BS1" : 0, "BS2" : 0, "BS3" : 0, "BS4" : 0, "BS5" : 0, "BS6" : 0} # счётчик опор каждого типа (диапазона)

# создаёт файл с результатами из 2 листов со списком опор и количеством опор каждого типа
tb = Workbook()
ws1 = tb.create_sheet("Supports")
ws2 = tb.create_sheet("Ranges")

col = 0 # счётчик опор
ws1.append(["#", "X", "Y", "H"]) # заголовок листа с опорами

# перебор координат от 0 до границ площади с заданным шагом
x = 0 
while x <= MAXX:
    ROW = []
    y = 0
    while y <= MAXY:
        z = Z(x, y)
        ROW.append({"X": x, "Y": y, "Z": z})
        ws1.append([col+1, x, y, z])
        for type, ran in RANGES.items():
            if ran[0] <= z < ran[1]:
                SUPTYPES[type]+=1
                break
        y+=STEPY
        col+=1
    SUPPS.append(ROW)
    x+=STEPX    

# вывод списка опор в консоль
for R in SUPPS:
    for val in R:
        print(val)
    print()
# print(SUPPS)

print(SUPTYPES) # вывод количества по диапазонам в консоль

# запись количества по диапазонам в Excel файл
for kd, vd in SUPTYPES.items():
    ws2.append([kd, str(RANGES[kd][0]) + '-' + str(RANGES[kd][1]),vd])
tb.remove(tb["Sheet"])
tb.save("outdata.xlsx")